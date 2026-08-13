from datetime import date, datetime, time
from io import BytesIO
from typing import BinaryIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from cairn_worker.errors import WorkerFailure
from cairn_worker.parsers import BlockKind
from cairn_worker.parsers.xlsx import XlsxParser
from openpyxl import Workbook

from apps.worker.tests.fixture_factory import empty_xlsx_fixture


def _save(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _structured_xlsx() -> bytes:
    workbook = Workbook()
    first = workbook.active
    assert first is not None
    first.title = "数据"
    first["A1"] = "name"
    first["C1"] = "value"
    first["B2"] = "formula label"
    first["D2"] = "=1+1"
    first["B4"] = True
    first["D4"] = 12.5
    first["A5"] = date(2026, 8, 13)
    first["B5"] = datetime.combine(date(2026, 8, 13), time(12, 34, 56))
    first["C5"] = time(12, 34, 56)
    first["D5"] = "#DIV/0!"
    first["D5"].data_type = "e"
    second = workbook.create_sheet("English")
    second["B3"] = "tail"
    return _save(workbook)


def _append_member(package: bytes, name: str, content: bytes) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(package), "r") as source, ZipFile(
        output, "w", compression=ZIP_DEFLATED
    ) as target:
        for member in source.infolist():
            target.writestr(member, source.read(member))
        target.writestr(name, content)
    return output.getvalue()


def _understate_first_worksheet_dimension(package: bytes) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(package), "r") as source, ZipFile(
        output, "w", compression=ZIP_DEFLATED
    ) as target:
        for member in source.infolist():
            data = source.read(member)
            if member.filename == "xl/worksheets/sheet1.xml":
                start = data.index(b"<dimension")
                end = data.index(b"/>", start) + 2
                data = data[:start] + b'<dimension ref="A1:A1"/>' + data[end:]
            target.writestr(member, data)
    return output.getvalue()


def _repeated_shared_string_xlsx(repetitions: int) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in range(1, repetitions + 1):
        worksheet.cell(row=row, column=1, value="repeat")
    package = _save(workbook)
    output = BytesIO()
    with ZipFile(BytesIO(package), "r") as source, ZipFile(
        output, "w", compression=ZIP_DEFLATED
    ) as target:
        for member in source.infolist():
            data = source.read(member)
            if member.filename == "xl/worksheets/sheet1.xml":
                data = data.replace(
                    b't="inlineStr"><is><t>repeat</t></is>', b't="s"><v>0</v>'
                )
            elif member.filename == "[Content_Types].xml":
                data = data.replace(
                    b"</Types>",
                    b'<Override PartName="/xl/sharedStrings.xml" '
                    b'ContentType="application/vnd.openxmlformats-officedocument.'
                    b'spreadsheetml.sharedStrings+xml"/></Types>',
                )
            target.writestr(member, data)
        target.writestr(
            "xl/sharedStrings.xml",
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            b'count="1" uniqueCount="1"><si><t>repeat</t></si></sst>',
        )
    return output.getvalue()


def test_xlsx_parser_emits_displayed_scalars_sparse_ranges_and_workbook_order() -> None:
    """Break caught: sparse row/column bounds and displayed scalar formatting must stay exact."""
    blocks = XlsxParser().parse(BytesIO(_structured_xlsx()))

    assert [block.kind for block in blocks] == [BlockKind.SHEET_ROWS, BlockKind.SHEET_ROWS]
    assert [block.text for block in blocks] == [
        (
            "name\t\tvalue\t\n\tformula label\t\t\n\tTRUE\t\t12.5\n"
            "2026-08-13T00:00:00\t2026-08-13T12:34:56\t12:34:56\t#DIV/0!"
        ),
        "tail",
    ]
    assert [block.locator.model_dump(by_alias=True) for block in blocks] == [
        {"type": "xlsx", "sheet": "数据", "cellRange": "A1:D5"},
        {"type": "xlsx", "sheet": "English", "cellRange": "B3:B3"},
    ]
    assert "=1+1" not in blocks[0].text


def test_xlsx_parser_groups_at_most_fifty_nonempty_source_rows() -> None:
    """Break caught: a sheet must not collapse unbounded source rows into one parser block."""
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Rows"
    for row in range(1, 52):
        worksheet.cell(row=row, column=1, value=f"row-{row}")

    blocks = XlsxParser().parse(BytesIO(_save(workbook)))

    assert [len(block.text.splitlines()) for block in blocks] == [50, 1]
    assert [block.locator.model_dump(by_alias=True)["cellRange"] for block in blocks] == [
        "A1:A50",
        "A51:A51",
    ]


def test_xlsx_parser_uses_safe_load_options_and_closes_on_success(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Break caught: formulas/links must stay inert and read-only workbooks must be closed."""
    from cairn_worker.parsers import xlsx as xlsx_parser

    real_load_workbook = xlsx_parser.load_workbook  # pyright: ignore[reportPrivateImportUsage]
    observed_kwargs: dict[str, bool] = {}
    close_count = [0]

    def tracking_load_workbook(
        filename: BinaryIO,
        *,
        read_only: bool = False,
        data_only: bool = False,
        keep_links: bool = True,
    ) -> Workbook:
        observed_kwargs.update(
            read_only=read_only,
            data_only=data_only,
            keep_links=keep_links,
        )
        workbook = real_load_workbook(
            filename,
            read_only=read_only,
            data_only=data_only,
            keep_links=keep_links,
        )
        original_close = workbook.close

        def tracking_close() -> None:
            close_count[0] += 1
            original_close()

        workbook.close = tracking_close
        return workbook

    monkeypatch.setattr(xlsx_parser, "load_workbook", tracking_load_workbook)
    XlsxParser().parse(BytesIO(_structured_xlsx()))

    assert observed_kwargs == {
        "read_only": True,
        "data_only": True,
        "keep_links": False,
    }
    assert close_count[0] == 1


def test_xlsx_parser_closes_an_open_workbook_when_row_iteration_fails(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Break caught: parser/library exceptions after open must still release the package handle."""
    from cairn_worker.parsers import xlsx as xlsx_parser

    real_load_workbook = xlsx_parser.load_workbook  # pyright: ignore[reportPrivateImportUsage]
    close_count = [0]

    def failing_load_workbook(
        filename: BinaryIO,
        *,
        read_only: bool = False,
        data_only: bool = False,
        keep_links: bool = True,
    ) -> Workbook:
        workbook = real_load_workbook(
            filename,
            read_only=read_only,
            data_only=data_only,
            keep_links=keep_links,
        )
        original_close = workbook.close

        def tracking_close() -> None:
            close_count[0] += 1
            original_close()

        workbook.close = tracking_close

        def fail_rows(*args: object, **kwargs: object) -> object:
            del args, kwargs
            raise RuntimeError("private worksheet failure")

        active = workbook.active
        assert active is not None
        active.iter_rows = fail_rows  # pyright: ignore[reportAttributeAccessIssue]
        return workbook

    monkeypatch.setattr(xlsx_parser, "load_workbook", failing_load_workbook)
    with pytest.raises(WorkerFailure) as caught:
        XlsxParser().parse(BytesIO(_structured_xlsx()))

    assert caught.value.code == "parser_failed"
    assert caught.value.safe_detail == "worker handler or parser failed"
    assert close_count[0] == 1


def test_xlsx_parser_rejects_malicious_dimensions_before_long_iteration(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Break caught: a tiny sparse package must not induce the full Excel grid as work."""
    from openpyxl.worksheet._read_only import (  # pyright: ignore[reportPrivateUsage]
        ReadOnlyWorksheet,
    )

    workbook = Workbook()
    active = workbook.active
    assert active is not None
    active["XFD1048576"] = "private distant cell"
    observed = {"iterated": False}

    def unexpected_iteration(*args: object, **kwargs: object) -> object:
        del args, kwargs
        observed["iterated"] = True
        raise AssertionError("malicious worksheet dimensions reached cell iteration")

    monkeypatch.setattr(ReadOnlyWorksheet, "iter_rows", unexpected_iteration)

    with pytest.raises(WorkerFailure) as caught:
        XlsxParser().parse(BytesIO(_save(workbook)))

    assert caught.value.code == "parser_failed"
    assert caught.value.retryable is False
    assert observed["iterated"] is False


@pytest.mark.parametrize(
    "content",
    [
        b"not a workbook",
        _append_member(empty_xlsx_fixture(), "xl/vbaProject.bin", b"private macro"),
    ],
    ids=("malformed", "macro"),
)
def test_xlsx_parser_rejects_malformed_or_macro_bearing_packages(content: bytes) -> None:
    """Break caught: malformed and executable spreadsheet content must fail safely."""
    with pytest.raises(WorkerFailure) as caught:
        XlsxParser().parse(BytesIO(content))

    assert caught.value.code == "parser_failed"
    assert caught.value.retryable is False
    assert caught.value.safe_detail == "worker handler or parser failed"


def test_xlsx_parser_rejects_an_empty_workbook() -> None:
    """Break caught: a valid workbook with only blank cells must not index an empty fact."""
    with pytest.raises(WorkerFailure) as caught:
        XlsxParser().parse(BytesIO(empty_xlsx_fixture()))
    assert caught.value.code == "no_extractable_text"


def test_xlsx_parser_never_silently_drops_cells_outside_understated_dimensions(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from cairn_worker.parsers import xlsx as xlsx_parser

    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["B3"] = "outside"
    content = _understate_first_worksheet_dimension(_save(workbook))
    close_count = [0]
    real_load = xlsx_parser.load_workbook  # pyright: ignore[reportPrivateImportUsage]

    def tracking_load(
        filename: BinaryIO,
        *,
        read_only: bool = False,
        data_only: bool = False,
        keep_links: bool = True,
    ) -> object:
        opened = real_load(
            filename,
            read_only=read_only,
            data_only=data_only,
            keep_links=keep_links,
        )
        original_close = opened.close

        def close() -> None:
            close_count[0] += 1
            original_close()

        opened.close = close
        return opened

    monkeypatch.setattr(xlsx_parser, "load_workbook", tracking_load)
    try:
        blocks = XlsxParser().parse(BytesIO(content))
    except WorkerFailure as failure:
        assert failure.code == "parser_failed"
    else:
        assert [block.text for block in blocks] == ["\toutside"]
        assert blocks[0].locator.model_dump(by_alias=True)["cellRange"] == "A1:B3"
    assert close_count == [1]


def test_xlsx_parser_rejects_repeated_shared_string_output_before_join(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from cairn_worker.parsers import xlsx as xlsx_parser

    joined = [False]

    def unexpected_block(*args: object, **kwargs: object) -> object:
        del args, kwargs
        joined[0] = True
        raise AssertionError("oversized repeated output reached block join")

    monkeypatch.setattr(xlsx_parser, "XLSX_MAX_OUTPUT_CHARACTERS", 20, raising=False)
    monkeypatch.setattr(xlsx_parser, "_sheet_block", unexpected_block)
    with pytest.raises(WorkerFailure) as caught:
        XlsxParser().parse(BytesIO(_repeated_shared_string_xlsx(19)))
    assert caught.value.code == "parser_failed"
    assert joined[0] is False
