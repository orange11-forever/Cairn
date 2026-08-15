import math
import re
from datetime import date, datetime, time
from io import BytesIO
from typing import Any, BinaryIO, cast
from xml.parsers import expat
from zipfile import ZipFile

from cairn_api.knowledge.schemas import XlsxLocator
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from cairn_worker.parsers import (
    BlockKind,
    DocumentParser,
    ParsedBlock,
    normalize_parser_text,
    read_parser_source,
)
from cairn_worker.parsers.limits import PARSER_SOURCE_MAX_BYTES, ensure_block_capacity
from cairn_worker.parsers.office_safety import validate_opc_package

XLSX_ROWS_PER_BLOCK = 50
XLSX_MAX_SHEETS = 1_000
XLSX_MAX_SOURCE_ROW = 100_000
XLSX_MAX_SOURCE_COLUMN = 4_096
XLSX_MAX_DIMENSION_CELLS = 2_000_000
XLSX_MAX_OUTPUT_BYTES = PARSER_SOURCE_MAX_BYTES
_XLSX_UTF8_LENGTH_CACHE_ENTRIES = 4_096
_CELL_REFERENCE = re.compile(r"\$?([A-Z]{1,3})\$?([1-9][0-9]*)\Z", re.IGNORECASE)


def _displayed_scalar(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError("non-finite spreadsheet number")
        return str(value)
    if isinstance(value, str):
        return value
    return None


def _sheet_bounds(worksheet: Worksheet) -> tuple[int, int, int, int]:
    declared = worksheet.calculate_dimension()
    declared_min_column, declared_min_row, declared_max_column, declared_max_row = (
        range_boundaries(declared)
    )
    if None in (
        declared_min_column,
        declared_min_row,
        declared_max_column,
        declared_max_row,
    ):
        raise ValueError("invalid worksheet dimension")
    assert declared_min_column is not None
    assert declared_min_row is not None
    assert declared_max_column is not None
    assert declared_max_row is not None
    if (
        declared_max_row > XLSX_MAX_SOURCE_ROW
        or declared_max_column > XLSX_MAX_SOURCE_COLUMN
        or (declared_max_row - declared_min_row + 1)
        * (declared_max_column - declared_min_column + 1)
        > XLSX_MAX_DIMENSION_CELLS
    ):
        raise ValueError("worksheet dimension exceeds parser work limit")
    read_only_worksheet = cast(Any, worksheet)
    read_only_worksheet.reset_dimensions()
    try:
        dimension = cast(str, read_only_worksheet.calculate_dimension(force=True))
    except UnboundLocalError:
        dimension = declared
    min_column, min_row, max_column, max_row = range_boundaries(dimension)
    if None in (min_column, min_row, max_column, max_row):
        raise ValueError("invalid worksheet dimension")
    assert min_column is not None
    assert min_row is not None
    assert max_column is not None
    assert max_row is not None
    min_column = min(min_column, declared_min_column)
    min_row = min(min_row, declared_min_row)
    max_column = max(max_column, declared_max_column)
    max_row = max(max_row, declared_max_row)
    if (
        max_row > XLSX_MAX_SOURCE_ROW
        or max_column > XLSX_MAX_SOURCE_COLUMN
        or (max_row - min_row + 1) * (max_column - min_column + 1)
        > XLSX_MAX_DIMENSION_CELLS
    ):
        raise ValueError("worksheet dimension exceeds parser work limit")
    return min_column, min_row, max_column, max_row


def _preflight_worksheet_coordinates(content: bytes) -> None:
    with ZipFile(BytesIO(content), "r") as package:
        worksheet_names = [
            name
            for name in package.namelist()
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        ]
        for name in worksheet_names:
            minimum_row = XLSX_MAX_SOURCE_ROW + 1
            minimum_column = XLSX_MAX_SOURCE_COLUMN + 1
            maximum_row = 0
            maximum_column = 0
            cells = 0
            parser = expat.ParserCreate()

            def start(element: str, attributes: dict[str, str]) -> None:
                nonlocal minimum_row, minimum_column, maximum_row, maximum_column, cells
                if element.rsplit(":", 1)[-1] != "c":
                    return
                reference = attributes.get("r")
                if reference is None:
                    raise ValueError("worksheet cell is missing a coordinate")
                match = _CELL_REFERENCE.fullmatch(reference)
                if match is None:
                    raise ValueError("worksheet cell has an invalid coordinate")
                column = 0
                for character in match.group(1).upper():
                    column = column * 26 + ord(character) - ord("A") + 1
                row = int(match.group(2))
                cells += 1
                minimum_row = min(minimum_row, row)
                minimum_column = min(minimum_column, column)
                maximum_row = max(maximum_row, row)
                maximum_column = max(maximum_column, column)
                if (
                    cells > XLSX_MAX_DIMENSION_CELLS
                    or row > XLSX_MAX_SOURCE_ROW
                    or column > XLSX_MAX_SOURCE_COLUMN
                    or (maximum_row - minimum_row + 1)
                    * (maximum_column - minimum_column + 1)
                    > XLSX_MAX_DIMENSION_CELLS
                ):
                    raise ValueError("worksheet coordinates exceed parser work limit")

            parser.StartElementHandler = start
            parser.Parse(package.read(name), True)


def _sheet_block(
    *,
    sheet_name: str,
    rows: list[tuple[int, dict[int, str]]],
    locator_floor: tuple[int, int] | None = None,
) -> ParsedBlock:
    min_column = min(column for _, values in rows for column in values)
    max_column = max(column for _, values in rows for column in values)
    row_start = rows[0][0]
    row_end = rows[-1][0]
    if locator_floor is not None:
        min_column = min(min_column, locator_floor[0])
        row_start = min(row_start, locator_floor[1])
    text = "\n".join(
        "\t".join(values.get(column, "") for column in range(min_column, max_column + 1))
        for _, values in rows
    )
    return ParsedBlock(
        kind=BlockKind.SHEET_ROWS,
        text=text,
        locator=XlsxLocator.model_validate(
            {
                "sheet": sheet_name,
                "cellRange": (
                    f"{get_column_letter(min_column)}{row_start}:"
                    f"{get_column_letter(max_column)}{row_end}"
                ),
            }
        ),
    )


def _utf8_byte_length(value: str, cache: dict[str, int]) -> int:
    cached = cache.get(value)
    if cached is None:
        cached = len(value.encode("utf-8"))
        if len(cache) >= _XLSX_UTF8_LENGTH_CACHE_ENTRIES:
            cache.clear()
        cache[value] = cached
    return cached


def _reserve_tsv_cell_bytes(
    value: str,
    *,
    column: int,
    previous_min_column: int | None,
    previous_max_column: int | None,
    row_count: int,
    row_started: bool,
    locator_min_column: int | None,
    consumed_bytes: int,
    utf8_lengths: dict[str, int],
    committed_pending_bytes: int,
    excluded_outer_newline_bytes: int,
) -> tuple[int, int, int]:
    if previous_min_column is None or previous_max_column is None:
        new_min = min(column, locator_min_column or column)
        new_max = column
        generated_bytes = new_max - new_min
    else:
        new_min = min(previous_min_column, column)
        new_max = max(previous_max_column, column)
        old_width = previous_max_column - previous_min_column + 1
        new_width = new_max - new_min + 1
        generated_bytes = (new_width - old_width) * (row_count + 1)
        if not row_started:
            generated_bytes += old_width - 1
            if row_count:
                generated_bytes += 1
    total = (
        consumed_bytes
        + generated_bytes
        + committed_pending_bytes
        + _utf8_byte_length(value, utf8_lengths)
        - excluded_outer_newline_bytes
    )
    if total > XLSX_MAX_OUTPUT_BYTES:
        raise ValueError("XLSX text exceeds parser output limit")
    return total, new_min, new_max


class XlsxParser(DocumentParser):
    def _parse(self, source: BinaryIO) -> list[ParsedBlock]:
        content = read_parser_source(source)
        validate_opc_package(content, required_member="xl/workbook.xml")
        _preflight_worksheet_coordinates(content)
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            if len(workbook.worksheets) > XLSX_MAX_SHEETS:
                raise ValueError("workbook sheet count exceeds parser work limit")
            blocks: list[ParsedBlock] = []
            output_bytes = 0
            utf8_lengths: dict[str, int] = {}
            for worksheet in workbook.worksheets:
                min_column, min_row, max_column, max_row = _sheet_bounds(worksheet)
                rows: list[tuple[int, dict[int, str]]] = []
                first_block = True
                block_min_column: int | None = None
                block_max_column: int | None = None
                block_leading_newline_bytes = 0
                block_trailing_newline_bytes = 0
                for row_number, cells in enumerate(
                    worksheet.iter_rows(
                        min_row=min_row,
                        max_row=max_row,
                        min_col=min_column,
                        max_col=max_column,
                    ),
                    start=min_row,
                ):
                    values: dict[int, str] = {}
                    row_started = False
                    for cell in cells:
                        displayed = _displayed_scalar(cell.value)
                        normalized = (
                            normalize_parser_text(displayed)
                            if displayed is not None
                            else None
                        )
                        if (
                            normalized is not None
                            and normalized.strip()
                            and isinstance(cell.column, int)
                        ):
                            leading_newlines = len(normalized) - len(
                                normalized.lstrip("\n")
                            )
                            trailing_newlines = len(normalized) - len(
                                normalized.rstrip("\n")
                            )
                            committed_pending = block_trailing_newline_bytes
                            block_trailing_newline_bytes = 0
                            if (
                                block_min_column is not None
                                and cell.column < block_min_column
                            ):
                                committed_pending += block_leading_newline_bytes
                                block_leading_newline_bytes = 0
                            excludes_leading = (
                                block_min_column is None
                                and min(cell.column, min_column if first_block else cell.column)
                                == cell.column
                            )
                            resulting_max = max(
                                block_max_column or cell.column, cell.column
                            )
                            excludes_trailing = cell.column == resulting_max
                            excluded_newlines = (
                                leading_newlines if excludes_leading else 0
                            ) + (trailing_newlines if excludes_trailing else 0)
                            output_bytes, block_min_column, block_max_column = (
                                _reserve_tsv_cell_bytes(
                                    normalized,
                                    column=cell.column,
                                    previous_min_column=block_min_column,
                                    previous_max_column=block_max_column,
                                    row_count=len(rows),
                                    row_started=row_started,
                                    locator_min_column=min_column if first_block else None,
                                    consumed_bytes=output_bytes,
                                    utf8_lengths=utf8_lengths,
                                    committed_pending_bytes=committed_pending,
                                    excluded_outer_newline_bytes=excluded_newlines,
                                )
                            )
                            if excludes_leading:
                                block_leading_newline_bytes = leading_newlines
                            if excludes_trailing:
                                block_trailing_newline_bytes = trailing_newlines
                            values[cell.column] = normalized
                            row_started = True
                    if not values:
                        continue
                    rows.append((row_number, values))
                    if len(rows) == XLSX_ROWS_PER_BLOCK:
                        ensure_block_capacity(len(blocks))
                        blocks.append(
                            _sheet_block(
                                sheet_name=worksheet.title,
                                rows=rows,
                                locator_floor=(min_column, min_row) if first_block else None,
                            )
                        )
                        first_block = False
                        rows = []
                        block_min_column = None
                        block_max_column = None
                        block_leading_newline_bytes = 0
                        block_trailing_newline_bytes = 0
                if rows:
                    ensure_block_capacity(len(blocks))
                    blocks.append(
                        _sheet_block(
                            sheet_name=worksheet.title,
                            rows=rows,
                            locator_floor=(min_column, min_row) if first_block else None,
                        )
                    )
            return blocks
        finally:
            workbook.close()


__all__ = ["XlsxParser"]
